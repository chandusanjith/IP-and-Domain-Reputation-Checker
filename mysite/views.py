from __future__ import print_function
from django.shortcuts import render
from django.shortcuts import redirect
from django.http import HttpResponse, HttpResponseNotFound, HttpResponseRedirect
import sys,re, smtplib
from mysite.models import ips, contacted, CountryData, IPData, Hashes, Ratings
from mysite.abuseipdb import abuseipdbChecker
from mysite.sans import sansChecker
from mysite.myIPwhois import IPWhoisChecker
from mysite.xforceIBM import myXForceChecker, myXForceHashChecker
from mysite.VirusTotal import VirusTotalChecker
import openpyxl
from django.contrib import messages
import xlwt
import datetime
from django.contrib.auth.models import User
import pydnsbl
import random
import string
from datetime import date
from pathos.multiprocessing import ProcessingPool as Pool
from threading import Thread
from django import db
import django
from django.views.static import serve
import os
from django.contrib.auth.models import User, auth
from django.contrib.auth import logout, login
from django.contrib.auth import logout as django_logout




def Rating(request):
  count = 1
  star1 = request.POST.get("stared1")
  star2 = request.POST.get("stared2")
  star3 = request.POST.get("stared3")
  star4 = request.POST.get("stared4")
  star5 = request.POST.get("stared5")
  print(star1)
  if star1 ==  "on":
    count = 1
  if star2 == "on":
    count = 2
  if star3 == "on":
    count = 3 
  if star4 == "on":
     count = 4
  if star5 == "on":
    count = 5
  a = Ratings(rating = count)
  a.save()
  messages.info(request, 'Thanks for the rating!!')
  return HttpResponseRedirect('/main/')

def DownloadResume(request):
    filepath = 'templates/RESUME.pdf' 
    return serve(request, os.path.basename(filepath),os.path.dirname(filepath))

def LoadPage(request):
      b = IPData.objects.all()
      a = CountryData.objects.all()
      total_rating_count = Ratings.objects.all().count()
      fivestar_count = Ratings.objects.filter(rating = 5).count()
      fourstar_count = Ratings.objects.filter(rating = 4).count()
      threestar_count = Ratings.objects.filter(rating = 3).count()
      twostar_count = Ratings.objects.filter(rating = 2).count()
      onestar_count = Ratings.objects.filter(rating = 1).count()
      avg_ratings = ((5 *fivestar_count) + (4 * fourstar_count) + (3*threestar_count) + (2*twostar_count) + (1*onestar_count))/total_rating_count
      rounded_average = round(avg_ratings)
      avg_ratings = round(avg_ratings, 1)
      context = {
       'total_rating_count':total_rating_count,
       'fivestar_count':fivestar_count,
        'fourstar_count':fourstar_count,
        'threestar_count':threestar_count,
        'twostar_count':twostar_count,
        'onestar_count':onestar_count,
        'avg_ratings':avg_ratings,
        'rounded_average':rounded_average,
      'cdata': a,
      'ipdata': b}
      return render(request, 'Main.html', context)

def LoadCheckIp(request):
       return render(request, 'checkip.html')

def About(request):
        return render(request, 'About.html')

def ContactDev(request):
        return render(request, 'Developer.html')


def addcontact(request):
      name = request.POST['namec']
      email = request.POST['email']
      message = request.POST['msg']
      mobileno = request.POST['mno']
      data = contacted(name = name, email = email, message = message, mobile = mobileno, dateofcontact = date.today())
      data.save()
      messages.info(request, 'Thanks, we will contact you soon ')
      return render(request, 'checkip.html')   

def Contact(request):
      return render(request, 'contactus.html')

def checksingleip(request):
      ip = request.POST['singleip']
      if ip == '':
          messages.info(request, 'Field is empty!!')
          return render(request, 'checkip.html')
      mySansPrint2 = ''
      myAbuseIPDBPrint3 = ''
      myXForcePrint4 = ''
      # regular expression for IP
      re_ip = re.compile('^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$')

      # If the IP format is valid:
      if re_ip.match(ip):
        # Call myIPwhois.py
        #part1 = IPWhoisChecker("https://www.abuseipdb.com/whois/" + ip)
        #part1 = part1[345:]
        part1 = checkippydnsbl(ip)      
        # Call sans.py
        part2 =  sansChecker(ip)
        # Call abuseipdb.py
        try:
          part3 =  abuseipdbChecker("https://www.abuseipdb.com/check/" + ip)
        except:
          part3 = "Local IP Possibly safe"      
        # Call xforceIBM.py    
        part4 = myXForceChecker("https://api.xforce.ibmcloud.com/ipr/" + ip)
        try:
          part5 = VirusTotalChecker(ip)
        except:
          part5 = "Virus Total Down, API Not responding!!!"
        part6 = IPWhoisChecker("https://www.abuseipdb.com/whois/" + ip)
        if part5 == "POSSIBLY SAFE" or part5 == "Virus Total Down, API Not responding!!!":
            res5 = "False"
        else:
            res5 = "True"

        if part1 == True or part1 == False:
          ip_checker = pydnsbl.DNSBLIpChecker()
          result=ip_checker.check(ip)
          resukt1 = result.detected_by
        
        if part3 != "Local IP Possibly safe" or "er":
          if not part3:
              req4 = "False"
          else:
              req4 = "True"
        else:
            part3 = " "
          
        
        if  "Malware" in part4[2]:
            req3 = "True"
        else:
            req3 = "False"               
        if (int(part2[0][13:])) > 0 :
            req2 = "True"
        else:
            req2 = "False"
        if part1 == True or req2 == "True" or req3 == "True" or req4 == "True" or res5 == "True":
            status = "09"
        else:
              status = "00"
        if  CountryData.objects.filter(country = part4[0][9:]).exists():
            c = CountryData.objects.filter(country = part4[0][9:])
            count = c[0].blklistcount
            count = count + 1
            country = CountryData.objects.filter(country = part4[0][9:]).update(blklistcount = count)
        else:
            c = CountryData(country = part4[0][9:],blklistcount = 1 )
            c.save()
        d = IPData.objects.filter(pk = 1)
        ipcount = d[0].ipcount + 1
        if status == "09":
            blacklistedip = d[0].blacklistedip + 1
            goodip = d[0].goodip
        else:
            blacklistedip = d[0].blacklistedip
            goodip = d[0].goodip + 1
        dat = IPData.objects.filter(pk = 1).update(ipcount = ipcount,blacklistedip = blacklistedip, goodip = goodip )

        context = {'part1': resukt1,
                'part2': part2,
                'part3':part3,
                'part4': part4,
                'part5':part5,
                'part6':part6,
                'status': status,
                'isip': ip,
                'dns': "00"}
        return render(request, 'checkip.html', context)    
      # If the input is a domain or other strings, let the website validate then ...
      else:
          #Call myIPwhois.py
          part1 =  IPWhoisChecker("https://www.abuseipdb.com/whois/" + ip)
          # Call sans.py
          # sans.sansChecker("https://isc.sans.edu/api/ip/" + ip)
          #part2 = sansChecker(ip)
          # Call abuseipdb.py
          part3 = abuseipdbChecker("https://www.abuseipdb.com/check/" + ip)
          part4 = myXForceChecker("https://api.xforce.ibmcloud.com/url/" + ip)
          #context = {
          #'part2': part1,
          #'part3':part3,
          #'part4':part4 }
          d = IPData.objects.filter(pk = 1)
          count = d[0].domaincount + 1
          dat = IPData.objects.filter(pk = 1).update(domaincount = count)
          context = {'part6': part1,
                'part2':part3,
                'part4': part4,
                'isip': ip,
                'dns': "01"}
          return render(request, 'checkip.html', context)    
          
def create_ref_code():
    return ''.join(random.choices(string.ascii_lowercase + string.digits, k=20))

def checkippydnsbl(ipaddress):
   ip_checker = pydnsbl.DNSBLIpChecker()
   result=ip_checker.check(ipaddress)
   resultfinal = result.blacklisted
   return resultfinal

def checkipchandu(ipaddress):
    result = sansChecker(ipaddress)
    return result

def checkIBM(ipaddress):
    result = myXForceChecker("https://api.xforce.ibmcloud.com/ipr/" + ipaddress)
    return result


def ReadBulk(request):
        ipss = request.POST['names']
        ip_list = ipss.split()
        if not ip_list:
          messages.info(request, 'Field is empty!!')
          return render(request, 'index.html')
        else:
            ref = create_ref_code()
            ip_length = len(ip_list)
            ref_list = list()
            for i in range(ip_length):
              ref_list.append(ref)
            p = Pool(10)
            p.map(check, ip_list, ref_list)
            result_to_display = ips.objects.filter(reference = ref_list[0])
            context = {'data_ip': result_to_display,
                    'reference': ref_list[0],
                    'button': 1}

            return render(request, 'index.html', context)
        

def ReadXl(request):
        if "GET" == request.method:
            return render(request, 'index.html', {'button': 0})
        else:
            excel_file = request.FILES["excel_file"]
            if not excel_file:
              messages.info(request, 'Please select a file!!')
              return render(request, 'index.html')
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(excel_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Sheet1"]
            excel_data = list()
            # iterating over the rows and
            # getting value from each cell in row
            ref = create_ref_code()
            row_data = list()
            ref_data = list()
            for row in worksheet.iter_rows():
                for cell in row:
                    row_data.append(str(cell.value))
                    ref_data.append(ref)
            p = Pool(10)
            p.map(check, row_data, ref_data)
            result_to_display = ips.objects.filter(reference = ref_data[0])
            context = {'data_ip': result_to_display,
                    'reference': ref_data[0],
                    'button': 1}
            return render(request, 'index.html', context)

def check(ip, ref):
  db.connections.close_all()
  status1 = checkippydnsbl(ip)
  status2 = checkipchandu(ip)
  status3 = checkIBM(ip)
  status4 = abuseipdbChecker("https://www.abuseipdb.com/check/" + ip)
  if not status4:
      req4 = "False"
  else:
      req4 = "True"
  if  "Malware" in status3[2]:
      req3 = "True"
  else:
      req3 = "False"               
  if (int(status2[0][13:])) > 0 :
      req2 = "True"
  else:
      req2 = "False"
  if status1 == True or req2 == "True" or req3 == "True" or req4 == "True":
      dbstatus = "BLACKLISTED"
  else:
      dbstatus = "POSSIBLY SAFE"
  if req4 == "True":
     abuse = status4[0]
  else:
     abuse = "NIL"
  if req3 == "True":
     ibm = status3
  else:
     ibm = "NIL"
  if req2 == "True":
     sans = status2
  else:
     sans = "NIL"  
  if status1 == True:
     ip_checker = pydnsbl.DNSBLIpChecker()
     result=ip_checker.check(ip)
     dnsbl = result.detected_by
  else:
     dnsbl = "NIL"
  if  CountryData.objects.filter(country = status3[0][9:]).exists():
        c = CountryData.objects.filter(country = status3[0][9:])
        count = c[0].blklistcount
        count = count + 1
        country = CountryData.objects.filter(country = status3[0][9:]).update(blklistcount = count)
  else:
        c = CountryData(country = status3[0][9:],blklistcount = 1 )
        c.save()
  d = IPData.objects.filter(pk = 1)
  ipcount = d[0].ipcount + 1
  if dbstatus == "BLACKLISTED":
        blacklistedip = d[0].blacklistedip + 1
        goodip = d[0].goodip
  else:
        blacklistedip = d[0].blacklistedip
        goodip = d[0].goodip + 1
  dat = IPData.objects.filter(pk = 1).update(ipcount = ipcount,blacklistedip = blacklistedip, goodip = goodip )
  a = ips(reference = ref, ipaddress = ip, status = dbstatus, remarks = abuse, Sans =sans, Pysbil =dnsbl , VirusTotal =   "virustotal", IbmXForce = ibm)
  a.save()
  

def export_users_xls(request, ref):
    filename = "ParameterLabs" + str(datetime.datetime.now()) + ".xls"
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ParameterLabs.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('parameter')
    # Sheet header, first row
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = ['IP Address', 'Status', 'AbuseIPDB Result', 'Sans Result', 'PYSBIL Result', 'IBM X-Force Result']
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)
    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()
    rows = ips.objects.filter(reference = ref).values_list('ipaddress','status','remarks','Sans','Pysbil','IbmXForce')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
    wb.save(response)
    return response


def FileHashSingle(request):
      if request.method == "GET":
          return render(request, 'checkfilehash.html')
      else:
         hashed = request.POST['singlehash']
         try:
            if hashed == '':
               messages.info(request, 'Field is empty!!')
               return render(request, 'checkfilehash.html')
            part1 = myXForceHashChecker("https://api.xforce.ibmcloud.com/malware/" + hashed)
            d = IPData.objects.filter(pk = 1)
            hashcount = d[0].hashcount + 1
            IPData.objects.filter(pk = 1).update(hashcount  = hashcount)
            context = {
           'family': part1[0][0],
           'type': part1[1],
           'risk': part1[2],
           'hashed': hashed,
            }
            return render(request, 'checkfilehash.html', context)
         except:
            messages.error(request, 'check input, error occured!!!!')
            return render(request, 'checkfilehash.html')
      

def HashBulkRead(request):
    hashes = request.POST['names']
    hashes_list = hashes.split()
    try:
        if not hashes_list:
            messages.info(request, 'Field is empty!!')
            return render(request, 'bulkfilehash.html')
        else:
            ref = create_ref_code()
            hash_length = len(hashes_list)
            ref_list = list()
            for i in range(hash_length):
                ref_list.append(ref)
                p = Pool(20)
                p.map(checkhash, hashes_list, ref_list)
                result_to_display = Hashes.objects.filter(reference = ref_list[0])
                context = {'data_ip': result_to_display,
                    'reference': ref_list[0],
                    'button': 1}
        return render(request, 'bulkfilehash.html', context)
    except:
            messages.info(request, 'check your input, error occured!!')
            return render(request, 'bulkfilehash.html')

def BulkHash(request):
        if "GET" == request.method:
            return render(request, 'bulkfilehash.html', {'button': 0})
        else:
            excel_file = request.FILES["excel_file"]
            if not excel_file:
              messages.info(request, 'Please select a file!!')
              return render(request, 'bulkfilehash.html')
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(excel_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Sheet1"]
            excel_data = list()
            # iterating over the rows and
            # getting value from each cell in row
            ref = create_ref_code()
            row_data = list()
            ref_data = list()
            for row in worksheet.iter_rows():
                for cell in row:
                    row_data.append(str(cell.value))
                    ref_data.append(ref)
            p = Pool(20)
            p.map(checkhash, row_data, ref_data)
            result_to_display = Hashes.objects.filter(reference = ref_data[0])
            context = {'data_ip': result_to_display,
                    'reference': ref_data[0],
                    'button': 1}

            return render(request, 'bulkfilehash.html', context)

def checkhash(hashed, ref):
    db.connections.close_all()
    part1 = myXForceHashChecker("https://api.xforce.ibmcloud.com/malware/" + hashed)
    a = Hashes(reference = ref, hashes = hashed, family =part1[0][0], type = part1[1], risk = part1[2] )
    a.save()
    d = IPData.objects.filter(pk = 1)
    hashcount = d[0].hashcount + 1
    IPData.objects.filter(pk = 1).update(hashcount  = hashcount)
    django.db.connection.close()

def downxlshash(request, ref):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ParameterLabs.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('parameter')
    # Sheet header, first row
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = ['File Hash', 'Risk', 'Type', 'Family']
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)
    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()
    rows = Hashes.objects.filter(reference = ref).values_list('hashes','risk','type','family')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
    wb.save(response)
    return response